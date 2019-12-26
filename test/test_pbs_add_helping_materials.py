import json
from helpers import *
from default import TestDefault
from mock import patch, MagicMock
from nose.tools import assert_raises
from requests import exceptions
from pbsexceptions import *
from openpyxl import Workbook

class TestPbsAddHelpingMaterial(TestDefault):

    """Test class for pbs add helping materials commands."""

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_json_with_info(self, auto_mock, find_mock):
        """Test add_helpingmaterials json with info field works."""
        auto_mock.return_value = (0, None)
        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.read.return_value = json.dumps([{'info': {'key': 'value'}}])

        pbclient = MagicMock()
        pbclient.create_helpingmaterial.return_value = {'id': 1, 'info': {'key': 'value'}}
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, 'json')
        assert res == '1 helping materials added to project: short_name', res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_json_from_filextension(self, auto_mock, find_mock):
        """Test add_helpingmaterials json without specifying file extension works."""
        auto_mock.return_value = (0, None)
        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.name = 'helpingmaterials.json'
        helpingmaterials.read.return_value = json.dumps([{'info': {'key': 'value'}}])

        pbclient = MagicMock()
        pbclient.create_helping_material.return_value = {'id': 1, 'info': {'key': 'value'}}
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, None)
        assert res == '1 helping materials added to project: short_name', res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_csv_with_info(self, auto_mock, find_mock):
        """Test add_helpingmaterials csv with info field works."""
        auto_mock.return_value = (0, None)
        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.read.return_value = "info, value\n, %s, 2" % json.dumps({'key':'value'})

        pbclient = MagicMock()
        pbclient.create_helping_material.return_value = {'id': 1, 'info': {'key': 'value'}}
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, 'csv')
        assert res == '1 helping materials added to project: short_name', res

    @patch('helpers.openpyxl.load_workbook')
    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_excel_with_info(self, auto_mock, find_mock, workbook_mock):
        """Test add_helpingmaterials excel with info field works."""
        auto_mock.return_value = (0, None)
        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()
        project.id = 1

        wb = Workbook()
        ws = wb.active

        headers = ['Column Name', 'foo']
        ws.append(headers)
        for row in range(2, 10):
            ws.append(['value', 'bar'])

        ws.append([None, None])
        ws.append([None, None])

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.read.return_value = wb

        workbook_mock.return_value = wb

        pbclient = MagicMock()
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, 'xlsx')
        self.config.pbclient.create_helpingmaterial.assert_called_with(project_id=find_mock().id,
                                                                       info={'column_name': 'value',
                                                                             'foo': 'bar'})
        assert res == '8 helping materials added to project: short_name', res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_csv_from_filextension(self, auto_mock,
                                                         find_mock):
        """Test add_helpingmaterials csv without specifying file extension works."""
        auto_mock.return_value = (0, None)

        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.name = 'helpingmaterials.csv'
        helpingmaterials.read.return_value = "info, value\n, %s, 2" % json.dumps({'key':'value'})

        pbclient = MagicMock()
        pbclient.create_helping_material.return_value = {'id': 1, 'info': {'key': 'value'}}
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, None)
        assert res == '1 helping materials added to project: short_name', res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_json_without_info(self, auto_mock, find_mock):
        """Test add_heping_materials json without info field works."""
        auto_mock.return_value = (0, None)

        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.read.return_value = json.dumps([{'key': 'value'}])

        pbclient = MagicMock()
        pbclient.create_helping_material.return_value = {'id': 1, 'info': {'key': 'value'}}
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, 'json')
        assert res == '1 helping materials added to project: short_name', res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_csv_without_info(self, auto_mock, find_mock):
        """Test add_helpingmaterials csv without info field works."""
        auto_mock.return_value = (0, None)

        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.read.return_value = "key, value\n, 1, 2"

        pbclient = MagicMock()
        pbclient.create_helping_material.return_value = {'id': 1, 'info': {'key': 'value'}}
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, 'csv')
        assert res == '1 helping materials added to project: short_name', res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_unknow_type_from_filextension(self,
                                                                 auto_mock, 
                                                                 find_mock):
        """Test add_helpingmaterials with unknown type from file extension works."""
        auto_mock.return_value = (0, None)
        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.name = 'helping.doc'
        helpingmaterials.read.return_value = "key, value\n, 1, 2"

        pbclient = MagicMock()
        pbclient.create_helping_material.return_value = {'id': 1, 'info': {'key': 'value'}}
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, None)
        assert res == ("Unknown format for the tasks file. Use json, csv, po or "
                      "properties."), res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_unknow_type(self, auto_mock, find_mock):
        """Test add_helpingmaterials with unknown type works."""
        auto_mock.return_value = (0, None)

        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.read.return_value = "key, value\n, 1, 2"

        pbclient = MagicMock()
        pbclient.create_helping_material.return_value = {'id': 1, 'info': {'key': 'value'}}
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, 'doc')
        assert res == ("Unknown format for the tasks file. Use json, csv, po or "
                      "properties."), res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_csv_connection_error(self, auto_mock, find_mock):
        """Test add_helpingmaterials csv connection error works."""
        auto_mock.return_value = (0, None)

        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.read.return_value = "key, value\n, 1, 2"

        pbclient = MagicMock()
        pbclient.create_helpingmaterial.side_effect = exceptions.ConnectionError
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, 'csv')
        assert res == "Connection Error! The server http://server is not responding", res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_material_json_connection_error(self, auto_mock, find_mock):
        """Test add_helpingmaterials json connection error works."""
        auto_mock.return_value = (0, None)

        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        tasks = MagicMock()
        tasks.read.return_value = json.dumps([{'key': 'value'}])

        pbclient = MagicMock()
        pbclient.create_helpingmaterial.side_effect = exceptions.ConnectionError
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, tasks, 'json')
        assert res == "Connection Error! The server http://server is not responding", res

    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helpingmaterial_another_error(self, auto_mock, find_mock):
        """Test add_tasks another error works."""
        auto_mock.return_value = (0, None)
        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()

        find_mock.return_value = project

        tasks = MagicMock()
        tasks.read.return_value = json.dumps([{'key': 'value'}])

        pbclient = MagicMock()
        pbclient.create_helpingmaterial.return_value = self.error
        self.config.pbclient = pbclient
        assert_raises(ProjectNotFound, _add_helpingmaterials, self.config,
                      tasks, 'json')

    def test_create_helping_material_info(self):
        """Test create_helping_material_info method works."""
        data = dict(info=dict(foo=1))
        helping_info, file_path = create_helping_material_info(data)
        assert helping_info == dict(foo=1)
        assert file_path is None

        data = dict(foo=1)
        helping_info, file_path = create_helping_material_info(data)
        assert helping_info == dict(foo=1)
        assert file_path is None

        data = dict(foo=1, file_path='file')
        helping_info, file_path = create_helping_material_info(data)
        assert helping_info == dict(foo=1)
        assert file_path == 'file'

    @patch('helpers.openpyxl.load_workbook')
    @patch('helpers.find_project_by_short_name')
    @patch('helpers.enable_auto_throttling')
    def test_add_helping_materials_excel_with_file(self, auto_mock, find_mock, workbook_mock):
        """Test add_helpingmaterials excel with file_path field works."""
        auto_mock.return_value = (0, None)

        project = MagicMock()
        project.name = 'name'
        project.short_name = 'short_name'
        project.description = 'description'
        project.info = dict()
        project.id = 1

        wb = Workbook()
        ws = wb.active

        headers = ['Column Name', 'foo', 'file Path']
        ws.append(headers)
        for row in range(2, 10):
            ws.append(['value', 'bar', '/tmp/file.jpg'])

        ws.append([None, None, None])
        ws.append([None, None, None])

        find_mock.return_value = project

        helpingmaterials = MagicMock()
        helpingmaterials.read.return_value = wb

        workbook_mock.return_value = wb

        pbclient = MagicMock()
        hm = MagicMock()
        hm.info = {'column_name': 'value', 'foo': 'bar'}
        hm.id = 1
        pbclient.create_helpingmaterial.return_value = hm
        self.config.pbclient = pbclient
        res = _add_helpingmaterials(self.config, helpingmaterials, 'xlsx')
        self.config.pbclient.create_helpingmaterial.assert_called_with(project_id=find_mock().id,
                                                                       file_path='/tmp/file.jpg',
                                                                       info={'column_name': 'value',
                                                                             'foo': 'bar'})
        self.config.pbclient.update_helping_material.assert_called_with(hm)

        assert res == '8 helping materials added to project: short_name', res

