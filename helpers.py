#!/usr/bin/env python
# -*- coding: utf-8 -*-

# This file is part of PyBOSSA.
#
# PyBOSSA is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# PyBOSSA is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with PyBOSSA.  If not, see <http://www.gnu.org/licenses/>.
"""
Helper functions for the pbs command line client.

This module exports the following methods:
    * find_project_by_short_name: return the project by short_name.
    * check_api_errors: check for API errors returned by a PyBossa server.
    * format_error: format error message.
    * format_json_task: format a CSV row into JSON.
"""
import re
import os
import csv
import json
import time
import click
import datetime
from io import StringIO
import polib
import openpyxl
import itertools
from requests import exceptions
import requests
from pbsexceptions import *
import logging
from watchdog.observers import Observer
from watchdog.events import PatternMatchingEventHandler
import calendar


__all__ = ['find_project_by_short_name', 'check_api_error',
           'format_error', 'format_json_task', '_create_project',
           '_update_project', '_add_tasks', 'create_task_info',
           '_delete_tasks', 'enable_auto_throttling',
           '_update_tasks_redundancy',
           '_update_project_watch', 'PbsHandler',
           '_update_task_presenter_bundle_js', 'row_empty',
           '_add_helpingmaterials', 'create_helping_material_info']


def _create_project(config):
    """Create a project in a PyBossa server."""
    try:
        response = config.pbclient.create_project(config.project['name'],
                                                  config.project['short_name'],
                                                  config.project['description'])
        check_api_error(response)
        return ("Project: %s created!" % config.project['short_name'])
    except exceptions.ConnectionError:
        return("Connection Error! The server %s is not responding" % config.server)
    except (ProjectNotFound, TaskNotFound):
        raise

def _update_project_watch(config, task_presenter, results,
                          long_description, tutorial):  # pragma: no cover
    """Update a project in a loop."""
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s - %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S')
    path = os.getcwd()
    event_handler = PbsHandler(config, task_presenter, results,
                               long_description, tutorial)
    observer = Observer()
    # We only want the current folder, not sub-folders
    observer.schedule(event_handler, path, recursive=False)
    observer.start()
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

def _update_task_presenter_bundle_js(project):
    """Append to template a distribution bundle js."""
    if os.path.isfile ('bundle.min.js'):
        with open('bundle.min.js') as f:
            js = f.read()
        project.info['task_presenter'] += "<script>\n%s\n</script>" % js
        return

    if os.path.isfile ('bundle.js'):
        with open('bundle.js') as f:
            js = f.read()
        project.info['task_presenter'] += "<script>\n%s\n</script>" % js

def _update_project(config, task_presenter, results,
                    long_description, tutorial):
    """Update a project."""
    try:
        # Get project
        project = find_project_by_short_name(config.project['short_name'],
                                             config.pbclient,
                                             config.all)
        # Update attributes
        project.name = config.project['name']
        project.short_name = config.project['short_name']
        project.description = config.project['description']
        # Update long_description
        with open(long_description, 'r') as f:
            project.long_description = f.read()
        # Update task presenter
        with open(task_presenter, 'r') as f:
            project.info['task_presenter'] = f.read()
        _update_task_presenter_bundle_js(project)
        # Update results
        with open(results, 'r') as f:
            project.info['results'] = f.read()
        # Update tutorial
        with open(tutorial, 'r') as f:
            project.info['tutorial'] = f.read()
        response = config.pbclient.update_project(project)
        check_api_error(response)
        return ("Project %s updated!" % config.project['short_name'])
    except exceptions.ConnectionError:
        return ("Connection Error! The server %s is not responding" % config.server)
    except ProjectNotFound:
        return ("Project not found! The project: %s is missing." \
                " Use the flag --all=1 to search in all the server " \
                % config.project['short_name'])
    except TaskNotFound:
        raise


def _load_data(data_file, data_type):
    """Load data from CSV, JSON, Excel, ..., formats."""
    if data_type is None:
        data_type = data_file.name.split('.')[-1]
    # Data list to process
    data = []
    # JSON type
    if data_type == 'json':
        raw_data = data_file.read()
        data = json.loads(raw_data)
        return data
    # CSV type
    elif data_type == 'csv':
        raw_data = data_file.read()
        csv_data = StringIO(raw_data)
        reader = csv.DictReader(csv_data, delimiter=',')
        for line in reader:
            data.append(line)
        return data
    elif data_type in ['xlsx', 'xlsm', 'xltx', 'xltm']:
        wb = openpyxl.load_workbook(data_file)
        ws = wb.active
        # First headers
        headers = []
        for row in ws.iter_rows(max_row=1):
            for cell in row:
                tmp = '_'.join(cell.value.split(" ")).lower()
                headers.append(tmp)
        # Simulate DictReader
        for row in ws.iter_rows(min_row=2):
            values = []
            for cell in row:
                values.append(cell.value)
            tmp = dict(list(zip(headers, values)))
            if len(values) == len(headers) and not row_empty(values):
                data.append(tmp)
        return data
    # PO type
    elif data_type == 'po':
        raw_data = data_file.read()
        po = polib.pofile(raw_data)
        for entry in po.untranslated_entries():
            data.append(entry.__dict__)
        return data
    # PROPERTIES type (used in Java and Firefox extensions)
    elif data_type == 'properties':
        raw_data = data_file.read()
        lines = raw_data.split('\n')
        for l in lines:
            if l:
                var_id, string = l.split('=')
                tmp = dict(var_id=var_id, string=string)
                data.append(tmp)
        return data
    else:
        return data


def _add_tasks(config, tasks_file, tasks_type, priority, redundancy):
    """Add tasks to a project."""
    try:
        project = find_project_by_short_name(config.project['short_name'],
                                             config.pbclient,
                                             config.all)
        data = _load_data(tasks_file, tasks_type)
        if len(data) == 0:
            return ("Unknown format for the tasks file. Use json, csv, po or "
                    "properties.")
        # If true, warn user
        # if sleep:  # pragma: no cover
        #     click.secho(msg, fg='yellow')
        # Show progress bar
        with click.progressbar(data, label="Adding Tasks") as pgbar:
            for d in pgbar:
                task_info = create_task_info(d)
                response = config.pbclient.create_task(project_id=project.id,
                                                       info=task_info,
                                                       n_answers=redundancy,
                                                       priority_0=priority)

                # Check if for the data we have to auto-throttle task creation
                sleep, msg = enable_auto_throttling(config, data)
                check_api_error(response)
                # If auto-throttling enabled, sleep for sleep seconds
                if sleep:  # pragma: no cover
                    time.sleep(sleep)
            return ("%s tasks added to project: %s" % (len(data),
                    config.project['short_name']))
    except exceptions.ConnectionError:
        return ("Connection Error! The server %s is not responding" % config.server)
    except (ProjectNotFound, TaskNotFound):
        raise


def _add_helpingmaterials(config, helping_file, helping_type):
    """Add helping materials to a project."""
    try:
        project = find_project_by_short_name(config.project['short_name'],
                                             config.pbclient,
                                             config.all)
        data = _load_data(helping_file, helping_type)
        if len(data) == 0:
            return ("Unknown format for the tasks file. Use json, csv, po or "
                    "properties.")
        # Show progress bar
        with click.progressbar(data, label="Adding Helping Materials") as pgbar:
            for d in pgbar:
                helping_info, file_path = create_helping_material_info(d)
                if file_path:
                    # Create first the media object
                    hm = config.pbclient.create_helpingmaterial(project_id=project.id,
                                                                info=helping_info,
                                                                file_path=file_path)
                    check_api_error(hm)

                    z = hm.info.copy()
                    z.update(helping_info)
                    hm.info = z
                    response = config.pbclient.update_helping_material(hm)
                    check_api_error(response)
                else:
                    response = config.pbclient.create_helpingmaterial(project_id=project.id,
                                                                      info=helping_info)
                check_api_error(response)
                # Check if for the data we have to auto-throttle task creation
                sleep, msg = enable_auto_throttling(config, data,
                                                    endpoint='/api/helpinmaterial')
                # If true, warn user
                if sleep:  # pragma: no cover
                    click.secho(msg, fg='yellow')
                # If auto-throttling enabled, sleep for sleep seconds
                if sleep:  # pragma: no cover
                    time.sleep(sleep)
            return ("%s helping materials added to project: %s" % (len(data),
                    config.project['short_name']))
    except exceptions.ConnectionError:
        return ("Connection Error! The server %s is not responding" % config.server)
    except (ProjectNotFound, TaskNotFound):
        raise



def _delete_tasks(config, task_id, limit=100, offset=0):
    """Delete tasks from a project."""
    try:
        project = find_project_by_short_name(config.project['short_name'],
                                             config.pbclient,
                                             config.all)
        if task_id:
            response = config.pbclient.delete_task(task_id)
            check_api_error(response)
            return "Task.id = %s and its associated task_runs have been deleted" % task_id
        else:
            limit = limit
            offset = offset
            tasks = config.pbclient.get_tasks(project.id, limit, offset)
            while len(tasks) > 0:
                for t in tasks:
                    response = config.pbclient.delete_task(t.id)
                    check_api_error(response)
                offset += limit
                tasks = config.pbclient.get_tasks(project.id, limit, offset)
            return "All tasks and task_runs have been deleted"
    except exceptions.ConnectionError:
        return ("Connection Error! The server %s is not responding" % config.server)
    except (ProjectNotFound, TaskNotFound):
        raise


def _update_tasks_redundancy(config, task_id, redundancy, limit=300, offset=0):
    """Update tasks redundancy from a project."""
    try:
        project = find_project_by_short_name(config.project['short_name'],
                                             config.pbclient,
                                             config.all)
        if task_id:
            response = config.pbclient.find_tasks(project.id, id=task_id)
            check_api_error(response)
            task = response[0]
            task.n_answers = redundancy
            response = config.pbclient.update_task(task)
            check_api_error(response)
            msg = "Task.id = %s redundancy has been updated to %s" % (task_id,
                                                                      redundancy)
            return msg
        else:
            limit = limit
            offset = offset
            tasks = config.pbclient.get_tasks(project.id, limit, offset)
            with click.progressbar(tasks, label="Updating Tasks") as pgbar:
                while len(tasks) > 0:
                    for t in pgbar:
                        t.n_answers = redundancy
                        response = config.pbclient.update_task(t)
                        check_api_error(response)
                        # Check if for the data we have to auto-throttle task update
                        sleep, msg = enable_auto_throttling(config, tasks)
                        # If auto-throttling enabled, sleep for sleep seconds
                        if sleep:  # pragma: no cover
                            time.sleep(sleep)
                    offset += limit
                    tasks = config.pbclient.get_tasks(project.id, limit, offset)
                return "All tasks redundancy have been updated"
    except exceptions.ConnectionError:
        return ("Connection Error! The server %s is not responding" % config.server)
    except (ProjectNotFound, TaskNotFound):
        raise


def find_project_by_short_name(short_name, pbclient, all=None):
    """Return project by short_name."""
    try:
        response = pbclient.find_project(short_name=short_name, all=all)
        check_api_error(response)
        if (len(response) == 0):
            msg = '%s not found! You can use the all=1 argument to \
                   search in all the server.'
            error = 'Project Not Found'
            raise ProjectNotFound(msg, error)
        return response[0]
    except exceptions.ConnectionError:
        raise
    except ProjectNotFound:
        raise


def check_api_error(api_response):
    print(api_response)
    """Check if returned API response contains an error."""
    if type(api_response) == dict and 'code' in api_response and api_response['code'] != 200:
            print(("Server response code: %s" % api_response['code']))
            print(("Server response: %s" % api_response))
            raise exceptions.HTTPError('Unexpected response', response=api_response)
    if type(api_response) == dict and (api_response.get('status') == 'failed'):
        if 'ProgrammingError' in api_response.get('exception_cls'):
            raise DatabaseError(message='PyBossa database error.',
                                error=api_response)
        if ('DBIntegrityError' in api_response.get('exception_cls') and
            'project' in api_response.get('target')):
            msg = 'PyBossa project already exists.'
            raise ProjectAlreadyExists(message=msg, error=api_response)
        if 'project' in api_response.get('target'):
            raise ProjectNotFound(message='PyBossa Project not found',
                                  error=api_response)
        if 'task' in api_response.get('target'):
            raise TaskNotFound(message='PyBossa Task not found',
                               error=api_response)
        else:
            print(("Server response: %s" % api_response))
            raise exceptions.HTTPError('Unexpected response', response=api_response)


def format_error(module, error):
    """Format the error for the given module."""
    logging.error(module)
    # Beautify JSON error
    print((error.message))
    print((json.dumps(error.error, sort_keys=True, indent=4, separators=(',', ': '))))
    exit(1)


def create_task_info(task):
    """Create task_info field."""
    task_info = None
    if task.get('info'):
        task_info = task['info']
    else:
        task_info = task
    return task_info


def create_helping_material_info(helping):
    """Create helping_material_info field."""
    helping_info = None
    file_path = None
    if helping.get('info'):
        helping_info = helping['info']
    else:
        helping_info = helping
    if helping_info.get('file_path'):
        file_path = helping_info.get('file_path')
        del helping_info['file_path']
    return helping_info, file_path


def enable_auto_throttling(config, data, limit=299, endpoint='/api/task'):
    "Return sleep time if more tasks than those " \
    "allowed by the server are requested."
    # Get header from server
    endpoint = config.server + endpoint
    headers = requests.head(endpoint).headers
    # Get limit
    server_limit = int(headers.get('X-RateLimit-Remaining', 0))
    limit = server_limit or limit
    # Get reset time
    reset_epoch = int(headers.get('X-RateLimit-Reset', 0))
    # Compute sleep time
    sleep = (reset_epoch -
             calendar.timegm(datetime.datetime.utcnow().utctimetuple()))
    msg = 'Warning: %s remaining hits to the endpoint.' \
          ' Auto-throttling enabled!' % limit
    # If we have less than 10 hits on the endpoint, sleep
    if limit <= 10:
        return (sleep, msg)
    else:
        return 0, None


def format_json_task(task_info):
    """Format task_info into JSON if applicable."""
    try:
        return json.loads(task_info)
    except:
        return task_info


def row_empty(row):
    """Check if all values in row are None."""
    for value in row:
        if value is not None:
            return False
    return True


class PbsHandler(PatternMatchingEventHandler):

    patterns = ['*/template.html', '*/tutorial.html',
                '*/long_description.md', '*/results.html',
                '*/bundle.js', '*/bundle.min.js']

    def __init__(self, config, task_presenter, results,
                 long_description, tutorial):
        super(PbsHandler, self).__init__()
        self.config = config
        self.task_presenter = task_presenter
        self.results = results
        self.long_description = long_description
        self.tutorial = tutorial

    def on_modified(self, event):
        what = 'directory' if event.is_directory else 'file'
        logging.info("Modified %s: %s", what, event.src_path)
        res = _update_project(self.config, self.task_presenter, self.results,
                              self.long_description, self.tutorial)
        logging.info(res)
